#!/usr/bin/env python3
"""Create a non-filing arithmetic control summary from normalized AEAT records."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

from aeat_book_common import (
    FIELDS_BY_BOOK,
    load_catalog,
    load_json_records,
    normalize_code,
    parse_decimal,
    record_payload,
    validate_record,
)
from validate_aeat_book import SHEET_ALIASES, detect_data_start, fields_for_sheet


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Resume bases y cuotas para conciliación previa al Pre303."
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("--period", choices=("1T", "2T", "3T", "4T"))
    parser.add_argument(
        "--profile",
        type=Path,
        help="Perfil JSON; aporta el periodo cuando no está repetido en cada fila.",
    )
    return parser.parse_args()


def amount(value: Any) -> Decimal:
    return parse_decimal(value) or Decimal("0")


def input_records(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() != ".xlsx":
        return load_json_records(path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    records: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_ALIASES:
            continue
        book = SHEET_ALIASES[sheet_name]
        sheet = workbook[sheet_name]
        expected = 40 if book == "BIENES-INVERSIÓN" else len(FIELDS_BY_BOOK[book])
        type_row, data_start = detect_data_start(sheet, expected)
        fields = fields_for_sheet(book, sheet, type_row)
        for row_number in range(data_start, sheet.max_row + 1):
            row = {
                field: sheet.cell(row_number, column).value
                for column, field in enumerate(fields, start=1)
            }
            if any(value not in (None, "") for value in row.values()):
                records.append({"libro_aeat": book, "registro_aeat": row})
    workbook.close()
    return records


def main() -> int:
    args = parse_args()
    profile = (
        json.loads(args.profile.read_text(encoding="utf-8"))
        if args.profile
        else {}
    )
    catalog = load_catalog()
    totals: dict[tuple[str, str, str, str], dict[str, Decimal | int]] = defaultdict(
        lambda: {
            "filas": 0,
            "base": Decimal("0"),
            "cuota_repercutida": Decimal("0"),
            "cuota_soportada": Decimal("0"),
            "cuota_deducible": Decimal("0"),
            "cuota_devengada_tecnica": Decimal("0"),
        }
    )
    skipped = 0
    excluded_by_period = 0
    invalid: list[str] = []
    for index, item in enumerate(input_records(args.input), start=1):
        book, row = record_payload(item)
        if profile:
            row, errors, _ = validate_record(book, row, profile, catalog)
            if errors:
                invalid.append(f"registro {index}: {'; '.join(errors)}")
                continue
        row_period = normalize_code(row.get("periodo"))
        if args.period and row_period != args.period:
            excluded_by_period += 1
            continue
        if book == "EXPEDIDAS":
            key = (
                book,
                str(row.get("clave_operacion") or ""),
                str(
                    row.get("calificacion_operacion")
                    or row.get("operacion_exenta")
                    or ""
                ),
                str(row.get("tipo_iva") or ""),
            )
        elif book == "RECIBIDAS":
            key = (
                book,
                str(row.get("clave_operacion_gasto") or ""),
                "ISP"
                if str(row.get("inversion_sujeto_pasivo") or "").upper()
                in {"S", "SI", "SÍ", "TRUE"}
                else "NO-ISP",
                str(row.get("tipo_iva") or ""),
            )
        else:
            skipped += 1
            continue
        try:
            increments = {
                "base": amount(row.get("base_imponible")),
                "cuota_repercutida": amount(row.get("cuota_iva_repercutida")),
                "cuota_soportada": amount(row.get("cuota_iva_soportada")),
                "cuota_deducible": amount(row.get("cuota_deducible")),
                "cuota_devengada_tecnica": Decimal("0"),
            }
            if book == "RECIBIDAS" and (
                str(row.get("inversion_sujeto_pasivo") or "").upper()
                in {"S", "SI", "SÍ", "TRUE"}
                or str(row.get("clave_operacion_gasto") or "") == "09"
            ):
                increments["cuota_devengada_tecnica"] = increments[
                    "cuota_soportada"
                ]
        except ValueError as error:  # importe malformado: no abortar la conciliación
            invalid.append(f"registro {index}: {error}")
            continue
        bucket = totals[key]
        bucket["filas"] = int(bucket["filas"]) + 1
        for field, increment in increments.items():
            bucket[field] = Decimal(bucket[field]) + increment

    groups = []
    for key, values in sorted(totals.items()):
        groups.append(
            {
                "libro": key[0],
                "clave": key[1],
                "tratamiento": key[2],
                "tipo_iva": key[3],
                **{
                    field: str(value.quantize(Decimal("0.01")))
                    if isinstance(value, Decimal)
                    else value
                    for field, value in values.items()
                },
            }
        )
    repercutida = sum(
        (Decimal(group["cuota_repercutida"]) for group in groups), Decimal("0")
    )
    deducible = sum(
        (Decimal(group["cuota_deducible"]) for group in groups), Decimal("0")
    )
    technical = sum(
        (Decimal(group["cuota_devengada_tecnica"]) for group in groups),
        Decimal("0"),
    )
    total_due = repercutida + technical
    result = {
        "period": args.period,
        "groups": groups,
        "control": {
            "cuota_repercutida": str(repercutida.quantize(Decimal("0.01"))),
            "cuota_devengada_tecnica": str(technical.quantize(Decimal("0.01"))),
            "cuota_devengada_control": str(total_due.quantize(Decimal("0.01"))),
            "cuota_deducible": str(deducible.quantize(Decimal("0.01"))),
            "diferencia_simple": str(
                (total_due - deducible).quantize(Decimal("0.01"))
            ),
        },
        "invalid_records": invalid,
        "skipped_investment_rows": skipped,
        "excluded_by_period": excluded_by_period,
        "warning": "Control aritmético: no es el modelo 303 ni aplica todas sus casillas, ajustes o datos censales.",
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
