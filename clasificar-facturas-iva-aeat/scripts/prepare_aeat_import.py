#!/usr/bin/env python3
"""Create an import-only AEAT workbook from a richer audit workbook."""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path

import openpyxl

from aeat_book_common import FIELDS_BY_BOOK
from validate_aeat_book import (
    SHEET_ALIASES,
    detect_data_start,
    fields_for_sheet,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Elimina hojas y columnas internas de un libro de auditoría para "
            "crear una copia de importación AEAT."
        )
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.is_file() or args.workbook.suffix.lower() != ".xlsx":
        raise FileNotFoundError("Se requiere un fichero XLSX.")
    workbook = openpyxl.load_workbook(args.workbook)
    removed_sheets = []
    removed_columns: dict[str, int] = {}
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in SHEET_ALIASES:
            workbook.remove(workbook[sheet_name])
            removed_sheets.append(sheet_name)
            continue
        book = SHEET_ALIASES[sheet_name]
        sheet = workbook[sheet_name]
        expected = 40 if book == "BIENES-INVERSIÓN" else len(FIELDS_BY_BOOK[book])
        type_row, _ = detect_data_start(sheet, expected)
        official_columns = len(fields_for_sheet(book, sheet, type_row))
        meaningful_extra = 0
        for column in range(official_columns + 1, sheet.max_column + 1):
            if any(
                sheet.cell(row, column).value not in (None, "")
                for row in range(1, sheet.max_row + 1)
            ):
                meaningful_extra += 1
        if sheet.max_column > official_columns:
            sheet.delete_cols(
                official_columns + 1, sheet.max_column - official_columns
            )
        if meaningful_extra:
            removed_columns[sheet_name] = meaningful_extra

    output = args.output
    if output.suffix.lower() != ".xlsx":
        output.mkdir(parents=True, exist_ok=True)
        output = output / f"{args.workbook.stem}_IMPORT_AEAT.xlsx"
    else:
        output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    try:
        os.chmod(output, 0o600)
    except OSError:
        pass  # el sistema de ficheros no admite permisos POSIX
    size = output.stat().st_size
    if size > 4 * 1024 * 1024:
        output.unlink(missing_ok=True)
        raise ValueError("La copia de importación supera 4 MB.")
    print(
        json.dumps(
            {
                "output": str(output.resolve()),
                "size_bytes": size,
                "removed_sheets": removed_sheets,
                "removed_audit_columns": removed_columns,
                "warning": (
                    "La copia aún debe superar el validador local y el Servicio "
                    "de validación de la AEAT."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
