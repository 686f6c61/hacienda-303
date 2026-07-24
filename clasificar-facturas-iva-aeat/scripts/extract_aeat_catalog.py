#!/usr/bin/env python3
"""Extract every AEAT dropdown code and activity mapping from official workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


SKILL_ROOT = Path(__file__).resolve().parent.parent
ASSET_ROOT = SKILL_ROOT / "assets" / "aeat-2026"
DEFAULT_OUTPUT = ASSET_ROOT / "codigos_aeat_2026.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrae el catálogo completo de códigos de las plantillas AEAT 2026."
    )
    parser.add_argument(
        "--persona-juridica",
        type=Path,
        default=ASSET_ROOT / "PLANTILLA_LIBROS_Pers_Juridicas.xlsx",
    )
    parser.add_argument(
        "--unificado",
        type=Path,
        default=ASSET_ROOT / "PLANTILLA_LIBROS_UNIFICADOS.xlsx",
    )
    parser.add_argument(
        "--epigrafes",
        type=Path,
        default=ASSET_ROOT / "Epigrafes_x_EEDD.xlsx",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def infer_code(display: str, explicit: Any) -> str:
    if explicit is not None and str(explicit).strip():
        return str(explicit).strip().replace(",", ".")
    match = re.match(r"^\s*([^-]+?)\s*-", display)
    return match.group(1).strip().replace(",", ".") if match else display.strip()


def strip_code(display: str, code: str) -> str:
    prefix = re.compile(rf"^\s*{re.escape(code)}\s*-\s*", re.IGNORECASE)
    return prefix.sub("", display, count=1).strip()


def extract_codes(path: Path) -> dict[str, list[dict[str, str]]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["CODIGO-LITERAL"]
    result: dict[str, list[dict[str, str]]] = defaultdict(list)
    seen: set[tuple[str, str, str]] = set()
    for row in sheet.iter_rows(values_only=True):
        category_raw, display_raw = row[0], row[1]
        if category_raw is None or display_raw is None:
            continue
        category = str(category_raw).strip()
        display = str(display_raw).strip()
        code = infer_code(display, row[3] if len(row) > 3 else None)
        literal = strip_code(display, code)
        key = (category, code, literal)
        if key in seen:
            continue
        seen.add(key)
        result[category].append(
            {"code": code, "literal": literal, "display": display}
        )
    workbook.close()
    return dict(sorted(result.items()))


def extract_activities(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Epígrafes"]
    result = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        code, group, description = row[1], row[2], row[3]
        if code is None or group is None:
            continue
        result.append(
            {
                "codigo_actividad": str(code).strip(),
                "grupo_epigrafe": str(group).strip(),
                "descripcion": str(description or "").strip(),
            }
        )
    workbook.close()
    return result


def main() -> int:
    args = parse_args()
    for path in (args.persona_juridica, args.unificado, args.epigrafes):
        if not path.is_file():
            raise FileNotFoundError(path)

    catalog: dict[str, Any] = {
        "metadata": {
            "publisher": "Agencia Estatal de Administración Tributaria",
            "version": "2026-06-08",
            "sources": {
                "persona_juridica": {
                    "file": args.persona_juridica.name,
                    "sha256": sha256_file(args.persona_juridica),
                },
                "unificado": {
                    "file": args.unificado.name,
                    "sha256": sha256_file(args.unificado),
                },
                "epigrafes": {
                    "file": args.epigrafes.name,
                    "sha256": sha256_file(args.epigrafes),
                },
            },
        },
        "profiles": {
            "persona_juridica": extract_codes(args.persona_juridica),
            "unificado_iva_irpf": extract_codes(args.unificado),
        },
        "actividades_economicas": extract_activities(args.epigrafes),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    counts = {
        profile: sum(len(items) for items in categories.values())
        for profile, categories in catalog["profiles"].items()
    }
    print(
        json.dumps(
            {
                "output": str(args.output.resolve()),
                "codes": counts,
                "activities": len(catalog["actividades_economicas"]),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
