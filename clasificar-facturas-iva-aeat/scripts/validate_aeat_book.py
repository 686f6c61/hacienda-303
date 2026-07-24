#!/usr/bin/env python3
"""Validate compact or full-layout AEAT VAT book workbooks locally."""

from __future__ import annotations

import argparse
import json
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

from aeat_book_common import FIELDS_BY_BOOK, load_catalog, validate_record


SHEET_ALIASES = {
    "EXPEDIDAS": "EXPEDIDAS",
    "EXPEDIDAS_INGRESOS": "EXPEDIDAS",
    "RECIBIDAS": "RECIBIDAS",
    "RECIBIDAS_GASTOS": "RECIBIDAS",
    "BIENES-INVERSIÓN": "BIENES-INVERSIÓN",
}
TYPE_MARKERS = ("DECIMAL", "ALFANUMÉRICO", "ALFANUMERICO", "FECHA(")
USAGE_MARKERS = {"IVA", "IRPF", "CONTROL INTERNO"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Valida localmente libros AEAT con encabezado compacto o diseño "
            "completo LSI/LSIJ."
        )
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--profile", required=True, type=Path)
    parser.add_argument(
        "--strict-import",
        action="store_true",
        help="Rechaza columnas de auditoría situadas después del diseño oficial.",
    )
    return parser.parse_args()


def is_empty_row(values: dict[str, Any]) -> bool:
    return not any(value not in (None, "") for value in values.values())


def decimal_or_zero(value: Any) -> Decimal:
    return Decimal("0") if value in (None, "") else Decimal(str(value))


def detect_type_row(sheet: Any, official_columns: int) -> int:
    best_row = 0
    best_count = 0
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        count = 0
        for column in range(1, official_columns + 1):
            value = sheet.cell(row_number, column).value
            text = str(value or "").upper()
            if any(marker in text for marker in TYPE_MARKERS):
                count += 1
        if count > best_count:
            best_row, best_count = row_number, count
    if best_count < max(5, official_columns // 2):
        raise ValueError("no se ha localizado la fila de tipos del encabezado")
    return best_row


def detect_data_start(sheet: Any, official_columns: int) -> tuple[int, int]:
    type_row = detect_type_row(sheet, official_columns)
    candidate = type_row + 1
    values = {
        str(sheet.cell(candidate, column).value or "").strip().upper()
        for column in range(1, sheet.max_column + 1)
    }
    values.discard("")
    if values and values.issubset(USAGE_MARKERS):
        candidate += 1
    return type_row, candidate


def fields_for_sheet(book: str, sheet: Any, type_row: int) -> list[str]:
    fields = FIELDS_BY_BOOK[book]
    if book != "BIENES-INVERSIÓN":
        return fields
    has_real_estate_columns = any(
        "REFERENCIA CATASTRAL"
        in str(sheet.cell(row, 41).value or "").upper()
        for row in range(1, type_row + 1)
    )
    return fields if has_real_estate_columns else fields[:39] + [fields[-1]]


def meaningful_extra_columns(sheet: Any, official_columns: int) -> int:
    extras = 0
    for column in range(official_columns + 1, sheet.max_column + 1):
        if any(
            sheet.cell(row, column).value not in (None, "")
            for row in range(1, sheet.max_row + 1)
        ):
            extras += 1
    return extras


def main() -> int:
    args = parse_args()
    profile = json.loads(args.profile.read_text(encoding="utf-8"))
    catalog = load_catalog()
    errors: list[str] = []
    warnings: list[str] = []
    if not args.workbook.is_file() or args.workbook.suffix.lower() != ".xlsx":
        raise FileNotFoundError("Se requiere un fichero XLSX.")
    size = args.workbook.stat().st_size
    if size > 4 * 1024 * 1024:
        errors.append("El fichero supera el límite oficial de 4 MB.")

    workbook = openpyxl.load_workbook(
        args.workbook, data_only=True, read_only=False
    )
    recognized = {
        sheet_name: SHEET_ALIASES[sheet_name]
        for sheet_name in workbook.sheetnames
        if sheet_name in SHEET_ALIASES
    }
    extra_sheets = [
        sheet_name
        for sheet_name in workbook.sheetnames
        if sheet_name not in SHEET_ALIASES
    ]
    if extra_sheets:
        message = f"hojas adicionales no oficiales: {', '.join(extra_sheets)}"
        if args.strict_import:
            errors.append(message)
        else:
            warnings.append(message)
    if "EXPEDIDAS" not in recognized.values():
        errors.append("Falta la hoja EXPEDIDAS o EXPEDIDAS_INGRESOS.")
    if "RECIBIDAS" not in recognized.values():
        errors.append("Falta la hoja RECIBIDAS o RECIBIDAS_GASTOS.")

    counts = {"EXPEDIDAS": 0, "RECIBIDAS": 0, "BIENES-INVERSIÓN": 0}
    layouts: dict[str, dict[str, int]] = {}
    duplicate_keys: dict[tuple[Any, ...], list[str]] = {}
    for sheet_name, book in recognized.items():
        sheet = workbook[sheet_name]
        expected_default = len(FIELDS_BY_BOOK[book])
        try:
            type_row, provisional_start = detect_data_start(
                sheet, 40 if book == "BIENES-INVERSIÓN" else expected_default
            )
        except ValueError as error:
            errors.append(f"{sheet_name}: {error}")
            continue
        fields = fields_for_sheet(book, sheet, type_row)
        official_columns = len(fields)
        _, data_start = detect_data_start(sheet, official_columns)
        layouts[sheet_name] = {
            "type_row": type_row,
            "data_start": data_start,
            "official_columns": official_columns,
        }
        extra_columns = meaningful_extra_columns(sheet, official_columns)
        if extra_columns > 0:
            message = (
                f"{sheet_name}: {extra_columns} columna(s) adicional(es) "
                "de auditoría fuera del diseño oficial"
            )
            if args.strict_import:
                errors.append(message)
            else:
                warnings.append(message)

        for row_number in range(data_start, sheet.max_row + 1):
            values = {
                field: sheet.cell(row_number, column).value
                for column, field in enumerate(fields, start=1)
            }
            if is_empty_row(values):
                continue
            counts[book] += 1
            prefix = f"{sheet_name}!{row_number}"
            try:
                normalized, row_errors, row_warnings = validate_record(
                    book, values, profile, catalog
                )
            except Exception as error:  # never abort a batch on one malformed row
                errors.append(f"{prefix}: validación interrumpida: {error}")
                continue
            errors.extend(f"{prefix}: {error}" for error in row_errors)
            warnings.extend(f"{prefix}: {warning}" for warning in row_warnings)
            if row_errors:
                continue

            if book == "EXPEDIDAS":
                key = (
                    book,
                    normalized.get("fecha_expedicion"),
                    normalized.get("serie"),
                    normalized.get("numero"),
                    normalized.get("tipo_iva"),
                    normalized.get("base_imponible"),
                    normalized.get("total_factura"),
                )
                if normalized.get("calificacion_operacion") == "S1":
                    expected = (
                        decimal_or_zero(normalized.get("base_imponible"))
                        + decimal_or_zero(normalized.get("cuota_iva_repercutida"))
                        + decimal_or_zero(
                            normalized.get("cuota_recargo_equivalencia")
                        )
                    )
                    total = decimal_or_zero(normalized.get("total_factura"))
                    if abs(expected - total) > Decimal("0.02"):
                        warnings.append(
                            f"{prefix}: subtotal no cuadra con base + IVA + recargo"
                        )
            elif book == "RECIBIDAS":
                key = (
                    book,
                    normalized.get("nif_expedidor"),
                    normalized.get("factura_expedidor_serie_numero"),
                    normalized.get("fecha_expedicion"),
                    normalized.get("tipo_iva"),
                    normalized.get("base_imponible"),
                    normalized.get("total_factura"),
                )
                technical_tax = (
                    normalized.get("inversion_sujeto_pasivo") == "S"
                    or normalized.get("clave_operacion_gasto") == "09"
                )
                if not technical_tax:
                    expected = (
                        decimal_or_zero(normalized.get("base_imponible"))
                        + decimal_or_zero(normalized.get("cuota_iva_soportada"))
                        + decimal_or_zero(
                            normalized.get("cuota_recargo_equivalencia")
                        )
                    )
                    total = decimal_or_zero(normalized.get("total_factura"))
                    if abs(expected - total) > Decimal("0.02"):
                        warnings.append(
                            f"{prefix}: subtotal no cuadra con base + IVA + recargo"
                        )
            else:
                key = (
                    book,
                    normalized.get("bien_identificador"),
                    normalized.get("fecha_inicio_utilizacion"),
                    normalized.get("referencia_externa"),
                )
            duplicate_keys.setdefault(key, []).append(prefix)

    for key, locations in duplicate_keys.items():
        if len(locations) > 1 and any(part not in (None, "") for part in key[1:]):
            warnings.append(f"posible duplicado exacto {key}: {', '.join(locations)}")
    workbook.close()
    result = {
        "valid": not errors,
        "file": str(args.workbook.resolve()),
        "size_bytes": size,
        "sheets": list(recognized),
        "extra_sheets": extra_sheets,
        "layouts": layouts,
        "rows": counts,
        "errors": errors,
        "warnings": warnings,
        "note": (
            "La validación local no sustituye el Servicio de validación de "
            "Libros Registro de la AEAT."
        ),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
