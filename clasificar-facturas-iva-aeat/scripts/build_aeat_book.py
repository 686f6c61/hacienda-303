#!/usr/bin/env python3
"""Build an AEAT 2026-compatible XLSX VAT book from reviewed JSON/JSONL rows."""

from __future__ import annotations

import argparse
import copy
import json
import os
import re
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from aeat_book_common import (
    ASSET_ROOT,
    FIELDS_BY_BOOK,
    excel_value,
    load_catalog,
    load_json_records,
    parse_decimal,
    record_payload,
    validate_record,
)


SOURCE_SHEETS = {
    "EXPEDIDAS": "EXPEDIDAS_INGRESOS",
    "RECIBIDAS": "RECIBIDAS_GASTOS",
    "BIENES-INVERSIÓN": "BIENES-INVERSIÓN",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera el libro registro XLSX en formato normalizado AEAT 2026."
    )
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--profile", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--include-investment-book", action="store_true")
    parser.add_argument(
        "--audit-output",
        type=Path,
        help="Crea una copia separada con columnas internas; no importarla en AEAT.",
    )
    parser.add_argument(
        "--fail-on-warnings",
        action="store_true",
        help="No genera el libro mientras existan avisos de revisión.",
    )
    return parser.parse_args()


def safe_filename_part(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")


def restrict_permissions(path: Path) -> None:
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass  # el sistema de ficheros no admite permisos POSIX


def expected_filename(profile: dict[str, Any]) -> str:
    type_code = (
        "T"
        if profile.get("catalog_profile") == "unificado_iva_irpf"
        else "C"
    )
    return (
        f"{profile['ejercicio']}{safe_filename_part(str(profile['nif']))}"
        f"{type_code}{safe_filename_part(str(profile['nombre']))}.xlsx"
    )


def copy_header_sheet(
    source: openpyxl.worksheet.worksheet.Worksheet,
    target: openpyxl.worksheet.worksheet.Worksheet,
    column_count: int,
) -> None:
    for row in range(1, 4):
        target.row_dimensions[row].height = source.row_dimensions[row].height
        for column in range(1, column_count + 1):
            source_cell = source.cell(row, column)
            target_cell = target.cell(row, column, source_cell.value)
            if source_cell.has_style:
                target_cell.alignment = copy.copy(source_cell.alignment)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format
    for column in range(1, column_count + 1):
        letter = get_column_letter(column)
        source_dimension = source.column_dimensions[letter]
        target_dimension = target.column_dimensions[letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
    for merged in source.merged_cells.ranges:
        if merged.max_row <= 3 and merged.max_col <= column_count:
            target.merge_cells(str(merged))
    target.freeze_panes = "A4"
    target.auto_filter.ref = f"A3:{get_column_letter(column_count)}3"
    target.sheet_view.showGridLines = source.sheet_view.showGridLines


def target_sheet_name(book: str, unified: bool) -> str:
    if unified and book == "EXPEDIDAS":
        return "EXPEDIDAS_INGRESOS"
    if unified and book == "RECIBIDAS":
        return "RECIBIDAS_GASTOS"
    return book


def audit_metadata(item: dict[str, Any]) -> dict[str, Any]:
    alerts = item.get("alertas") or []
    return {
        "archivo_origen": item.get("archivo"),
        "sha256_origen": item.get("archivo_sha256") or item.get("sha256"),
        "operacion_id": item.get("operacion_id"),
        "case_id_aeat": item.get("case_id"),
        "confianza": item.get("confianza"),
        "clasificacion_iva": item.get("tratamiento"),
        "alertas": "; ".join(map(str, alerts)) if isinstance(alerts, list) else alerts,
        "factura_total_documento": item.get("factura_total_documento"),
    }


def reference_from_audit(audit: dict[str, Any]) -> str | None:
    for field in ("operacion_id", "sha256_origen", "archivo_origen"):
        value = audit.get(field)
        if value not in (None, ""):
            return str(value)[:40]
    return None


def save_audit_workbook(
    import_path: Path,
    audit_path: Path,
    rows: dict[str, list[dict[str, Any]]],
    unified: bool,
) -> Path:
    audit_book = openpyxl.load_workbook(import_path)
    headers = [
        ("Archivo origen", "Alfanumérico"),
        ("SHA-256 origen", "Alfanumérico"),
        ("Operación ID", "Alfanumérico"),
        ("Case ID AEAT", "Alfanumérico"),
        ("Confianza", "Alfanumérico"),
        ("Clasificación IVA", "Alfanumérico"),
        ("Alertas", "Alfanumérico"),
        ("Total factura documental", "Decimal(12,2)"),
    ]
    for book, metadata_rows in rows.items():
        sheet_name = target_sheet_name(book, unified)
        if sheet_name not in audit_book.sheetnames:
            continue
        sheet = audit_book[sheet_name]
        start = len(FIELDS_BY_BOOK[book]) + 1
        for offset, (label, value_type) in enumerate(headers):
            column = start + offset
            sheet.cell(1, column, f"{label} — NO IMPORTAR EN AEAT")
            sheet.cell(3, column, value_type)
            sheet.column_dimensions[get_column_letter(column)].width = 24
        for row_number, metadata in enumerate(metadata_rows, start=4):
            for offset, key in enumerate(
                (
                    "archivo_origen",
                    "sha256_origen",
                    "operacion_id",
                    "case_id_aeat",
                    "confianza",
                    "clasificacion_iva",
                    "alertas",
                    "factura_total_documento",
                )
            ):
                sheet.cell(row_number, start + offset, metadata.get(key))
    if audit_path.suffix.lower() != ".xlsx":
        audit_path.mkdir(parents=True, exist_ok=True)
        audit_path = audit_path / f"{import_path.stem}_AUDITORIA.xlsx"
    else:
        audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_book.save(audit_path)
    audit_book.close()
    restrict_permissions(audit_path)
    return audit_path


def main() -> int:
    args = parse_args()
    profile = json.loads(args.profile.read_text(encoding="utf-8"))
    for field in ("ejercicio", "nif", "nombre"):
        if profile.get(field) in (None, ""):
            raise ValueError(f"perfil: falta {field}")
    if profile.get("sii"):
        raise ValueError("Este exportador Pre303/LSI no es para contribuyentes SII.")
    if profile.get("periodicidad", "trimestral") != "trimestral":
        raise ValueError("Pre303/LSI requiere periodicidad trimestral.")

    catalog = load_catalog()
    raw_records = load_json_records(args.input)
    rows: dict[str, list[dict[str, Any]]] = {
        "EXPEDIDAS": [],
        "RECIBIDAS": [],
        "BIENES-INVERSIÓN": [],
    }
    audit_rows: dict[str, list[dict[str, Any]]] = {
        "EXPEDIDAS": [],
        "RECIBIDAS": [],
        "BIENES-INVERSIÓN": [],
    }
    all_errors: list[str] = []
    all_warnings: list[str] = []
    invoice_totals: dict[str, dict[str, Any]] = {}
    for index, item in enumerate(raw_records, start=1):
        if item.get("estado") and item["estado"] != "clasificacion_concluida":
            all_errors.append(
                f"registro {index}: no puede exportarse con estado {item['estado']!r}"
            )
            continue
        book, values = record_payload(item)
        values = dict(values)
        audit = audit_metadata(item)
        values.setdefault("referencia_externa", reference_from_audit(audit))
        try:
            normalized, errors, warnings = validate_record(
                book, values, profile, catalog
            )
        except ValueError as error:  # nunca abortar el lote por una fila malformada
            all_errors.append(f"registro {index}: validación interrumpida: {error}")
            continue
        all_errors.extend(f"registro {index}: {error}" for error in errors)
        all_warnings.extend(f"registro {index}: {warning}" for warning in warnings)
        if not errors:
            rows[book].append(normalized)
            audit_rows[book].append(audit)
            try:
                documented_total = parse_decimal(
                    audit.get("factura_total_documento")
                )
            except ValueError as error:
                all_errors.append(
                    f"registro {index}: factura_total_documento: {error}"
                )
                continue
            invoice_identity = str(
                audit.get("archivo_origen")
                or str(audit.get("operacion_id") or "").split("#", 1)[0]
            )
            if documented_total is not None and invoice_identity:
                group_key = f"{book}:{invoice_identity}"
                group = invoice_totals.setdefault(
                    group_key,
                    {
                        "expected": documented_total,
                        "subtotal": Decimal("0"),
                        "lineas": 0,
                    },
                )
                if group["expected"] != documented_total:
                    all_errors.append(
                        f"{group_key}: factura_total_documento incoherente entre líneas"
                    )
                group["subtotal"] += normalized.get("total_factura") or Decimal("0")
                group["lineas"] += 1
    for group_key, amounts in invoice_totals.items():
        # Cada línea redondea a céntimos (±0,005), así que la tolerancia crece
        # con el número de líneas de la factura.
        tolerance = max(Decimal("0.02"), Decimal("0.01") * amounts["lineas"])
        if abs(amounts["subtotal"] - amounts["expected"]) > tolerance:
            all_errors.append(
                f"{group_key}: suma de subtotales {amounts['subtotal']} no coincide "
                f"con total documental {amounts['expected']} "
                f"(tolerancia {tolerance})"
            )
    if all_errors:
        print(
            json.dumps(
                {
                    "valid": False,
                    "errors": all_errors,
                    "warnings": all_warnings,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 1
    if args.fail_on_warnings and all_warnings:
        print(
            json.dumps(
                {
                    "valid": False,
                    "errors": ["existen avisos y se solicitó --fail-on-warnings"],
                    "warnings": all_warnings,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 1

    unified = profile.get("catalog_profile") == "unificado_iva_irpf"
    template_name = (
        "PLANTILLA_LIBROS_UNIFICADOS.xlsx"
        if unified
        else "PLANTILLA_LIBROS_Pers_Juridicas.xlsx"
    )
    template = openpyxl.load_workbook(
        ASSET_ROOT / template_name, data_only=False, read_only=False
    )
    output_book = Workbook()
    output_book.remove(output_book.active)
    included_books = ["EXPEDIDAS", "RECIBIDAS"]
    if args.include_investment_book or rows["BIENES-INVERSIÓN"]:
        included_books.append("BIENES-INVERSIÓN")

    for book in included_books:
        fields = FIELDS_BY_BOOK[book]
        source_sheet = template[SOURCE_SHEETS[book]]
        sheet = output_book.create_sheet(target_sheet_name(book, unified))
        copy_header_sheet(source_sheet, sheet, len(fields))
        style_source_row = 4
        for row_index, values in enumerate(rows[book], start=4):
            for column, field in enumerate(fields, start=1):
                source_cell = source_sheet.cell(style_source_row, column)
                target_cell = sheet.cell(
                    row_index, column, excel_value(field, values.get(field))
                )
                if source_cell.has_style:
                    target_cell.alignment = copy.copy(source_cell.alignment)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.number_format = source_cell.number_format
                if field.startswith("fecha_") or field.endswith("_fecha"):
                    target_cell.number_format = "dd/mm/yyyy"
                elif isinstance(target_cell.value, float):
                    target_cell.number_format = "0.00"
        sheet.auto_filter.ref = (
            f"A3:{get_column_letter(len(fields))}{max(3, sheet.max_row)}"
        )

    output_book.calculation.fullCalcOnLoad = True
    output_book.calculation.forceFullCalc = True
    output_book.calculation.calcMode = "auto"
    template.close()

    output_path = args.output
    if output_path.suffix.lower() != ".xlsx":
        output_path.mkdir(parents=True, exist_ok=True)
        output_path = output_path / expected_filename(profile)
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    output_book.save(output_path)
    output_book.close()
    restrict_permissions(output_path)
    size = output_path.stat().st_size
    if size > 4 * 1024 * 1024:
        output_path.unlink(missing_ok=True)
        raise ValueError("El XLSX supera el límite oficial de 4 MB.")

    audit_path = None
    if args.audit_output:
        audit_path = save_audit_workbook(
            output_path, args.audit_output, audit_rows, unified
        )

    print(
        json.dumps(
            {
                "valid": True,
                "output": str(output_path.resolve()),
                "expected_filename": expected_filename(profile),
                "size_bytes": size,
                "audit_output": str(audit_path.resolve()) if audit_path else None,
                "rows": {book: len(rows[book]) for book in included_books},
                "warnings": all_warnings,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
